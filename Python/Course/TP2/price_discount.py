import openpyxl

def recalculate_price(path: str = './transactions.xlsx'):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws['D1'] = 'corrected'
    for i in range(2, 5):
        cell_C = ws[f'C{i}'].value.split(' ')[0].replace(',', '.')
        ws[f'D{i}'] = round(float(cell_C) * 0.9, 2)
    wb.save(path)
    wb.close()

recalculate_price()